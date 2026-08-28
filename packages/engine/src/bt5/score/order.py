"""The vendor order file -- the last artefact between a design and real DNA.

The IDT eBlocks 96-well plate upload template was inspected directly rather than
described from memory, and it is simpler than it sounds:

  - the format is .xlsx, not .csv, so this writes a real workbook
  - the SHEET NAME is the plate name; plate identity travels as sheet metadata,
    not as a column
  - exactly three columns: Well Position, Name, Sequence
  - 96 rows pre-filled with well positions in row-major order, A1..A12, B1..B12,
    ... H1..H12, whether or not every well is used

A plain Name,Sequence CSV covers tube orders and other vendors. Vendor profiles
carry a last_verified date because these templates drift.

Two things this module refuses to do, both because the failure is expensive and
silent. It will not write a sequence that is not bare ACGT -- a lowercase base
or a stray IUPAC N reaches the synthesiser as an order, not as a question. And
it will not write two wells with the same Name: the whole reason the design hash
goes on the tube label is that two different sequences under one name is how a
lab ends up with two tubes and an irreproducible result.
"""

from __future__ import annotations

import csv
import io
import re
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import IO

from bt5.core.types import DNA_ALPHABET

#: The plate the IDT template ships: 8 rows A-H, 12 columns, row-major.
PLATE_ROWS: tuple[str, ...] = tuple("ABCDEFGH")
PLATE_COLUMNS: tuple[int, ...] = tuple(range(1, 13))
PLATE_SIZE = len(PLATE_ROWS) * len(PLATE_COLUMNS)

IDT_HEADERS: tuple[str, str, str] = ("Well Position", "Name", "Sequence")

#: The template's own default sheet name, which the user renames to their plate.
DEFAULT_PLATE_NAME = "Plate Name"

#: openpyxl raises on these in a sheet title, and Excel silently mangles a title
#: over 31 characters -- both of which turn "plate identity" into "some other
#: plate's identity" at the vendor's end.
_ILLEGAL_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
MAX_SHEET_TITLE = 31

VENDOR_LAST_VERIFIED = "2026-08-28"


class OrderError(ValueError):
    """The order file cannot be written as asked."""


def wells(count: int = PLATE_SIZE) -> tuple[str, ...]:
    """Well positions in the template's row-major order."""
    if not 0 <= count <= PLATE_SIZE:
        raise OrderError(f"a plate holds {PLATE_SIZE} wells, asked for {count}")
    return tuple(f"{row}{col}" for row in PLATE_ROWS for col in PLATE_COLUMNS)[:count]


@dataclass(frozen=True, slots=True)
class OrderEntry:
    """One well: what to synthesise and what to call it.

    `name` is expected to carry the construct name AND the short design hash, so
    the tube label traces back to the run that produced it. Not enforced here --
    this module cannot know what the caller's naming scheme is -- but
    `entry_name` builds the form BT5 uses.
    """

    name: str
    sequence: str

    def __post_init__(self) -> None:
        if not self.name.strip():
            raise OrderError("an order entry needs a name; the tube label is the only handle")
        if not self.sequence:
            raise OrderError(f"{self.name}: empty sequence")
        bad = sorted(set(self.sequence) - DNA_ALPHABET)
        if bad:
            raise OrderError(
                f"{self.name}: order sequences must be bare uppercase ACGT, found {bad}. "
                f"A lowercase base or an IUPAC ambiguity code reaches the "
                f"synthesiser as an order rather than as a question."
            )

    @property
    def length_bp(self) -> int:
        return len(self.sequence)


def entry_name(construct_name: str, design_hash: str) -> str:
    """The tube label BT5 writes: name plus the short content hash.

    Two runs producing two different sequences under one name is how a lab ends
    up with two tubes and an irreproducible result, so the hash is not optional
    decoration.
    """
    if not design_hash.strip():
        raise OrderError("a tube label without its design hash is not traceable")
    return f"{construct_name}_{design_hash}"


def _checked(entries: Sequence[OrderEntry]) -> tuple[OrderEntry, ...]:
    if not entries:
        raise OrderError("nothing to order")
    names = [e.name for e in entries]
    duplicated = sorted({n for n in names if names.count(n) > 1})
    if duplicated:
        raise OrderError(
            f"duplicate order names {duplicated}: two wells under one label is "
            f"exactly what the design hash on the label exists to prevent"
        )
    return tuple(entries)


def plates(entries: Sequence[OrderEntry]) -> tuple[tuple[OrderEntry, ...], ...]:
    """Split an order into plate-sized chunks, in order.

    Returned rather than written, so the caller names each plate. Silently
    truncating at 96 would drop designs; raising would make a 100-design run
    unorderable for no reason.
    """
    checked = _checked(entries)
    return tuple(tuple(checked[i : i + PLATE_SIZE]) for i in range(0, len(checked), PLATE_SIZE))


def _sheet_title(plate_name: str) -> str:
    title = plate_name.strip() or DEFAULT_PLATE_NAME
    if _ILLEGAL_SHEET_CHARS.search(title):
        raise OrderError(
            f"plate name {plate_name!r} contains a character Excel forbids in a "
            f"sheet title ([ ] : * ? / \\). The sheet name IS the plate name in "
            f"this template, so it cannot be silently rewritten."
        )
    if len(title) > MAX_SHEET_TITLE:
        raise OrderError(
            f"plate name {plate_name!r} is {len(title)} characters; Excel truncates "
            f"a sheet title at {MAX_SHEET_TITLE}, which would send the vendor a "
            f"different plate name than the one on your bench"
        )
    return title


def write_idt_plate(
    entries: Sequence[OrderEntry],
    target: str | Path | IO[bytes] | None = None,
    *,
    plate_name: str = DEFAULT_PLATE_NAME,
) -> bytes:
    """Write one IDT eBlocks 96-well plate upload workbook.

    Returns the workbook bytes, and writes them to `target` when given. All 96
    well positions are emitted whether or not they are used, matching the
    template; unused wells carry an empty Name and Sequence.
    """
    checked = _checked(entries)
    if len(checked) > PLATE_SIZE:
        raise OrderError(
            f"{len(checked)} designs do not fit one {PLATE_SIZE}-well plate; "
            f"split with plates() and name each one"
        )

    try:
        from openpyxl import Workbook  # type: ignore[import-untyped]
    except ModuleNotFoundError as exc:  # pragma: no cover - depends on the extra
        raise OrderError(
            "the IDT plate format is a real .xlsx workbook and needs openpyxl: "
            "install the 'export' extra"
        ) from exc

    book = Workbook()
    sheet = book.active
    sheet.title = _sheet_title(plate_name)
    sheet.append(list(IDT_HEADERS))
    for index, position in enumerate(wells()):
        entry = checked[index] if index < len(checked) else None
        sheet.append(
            [position, entry.name, entry.sequence] if entry is not None else [position, "", ""]
        )

    buffer = io.BytesIO()
    book.save(buffer)
    data = buffer.getvalue()
    if target is not None:
        if isinstance(target, str | Path):
            Path(target).write_bytes(data)
        else:
            target.write(data)
    return data


def write_csv(entries: Sequence[OrderEntry], target: str | Path | IO[str] | None = None) -> str:
    """A plain Name,Sequence CSV, for tube orders and other vendors.

    `\\r\\n` line endings, because that is what a spreadsheet round-trips
    without complaint and what the csv module's own dialect specifies.
    """
    checked = _checked(entries)
    handle = io.StringIO()
    writer = csv.writer(handle)
    writer.writerow(["Name", "Sequence"])
    for entry in checked:
        writer.writerow([entry.name, entry.sequence])
    text = handle.getvalue()
    if target is not None:
        if isinstance(target, str | Path):
            Path(target).write_text(text, newline="")
        else:
            target.write(text)
    return text


def read_plate(source: str | Path | IO[bytes]) -> tuple[tuple[str, str, str], ...]:
    """Read back a written plate as (well, name, sequence) rows, header excluded.

    Here so the writer can be tested against a real workbook rather than against
    its own idea of one -- the failure this guards is a file that opens cleanly
    in openpyxl and is rejected by the vendor's uploader.
    """
    from openpyxl import load_workbook

    book = load_workbook(source)
    sheet = book.active
    rows = list(sheet.iter_rows(values_only=True))
    return tuple((str(r[0]), str(r[1] or ""), str(r[2] or "")) for r in rows[1:])


def sheet_name_of(source: str | Path | IO[bytes]) -> str:
    """The plate name, which in this template lives only in the sheet title."""
    from openpyxl import load_workbook

    return str(load_workbook(source).sheetnames[0])


def order_entries(
    designs: Iterable[tuple[str, str, str]],
) -> tuple[OrderEntry, ...]:
    """Build entries from (construct_name, design_hash, sequence) triples."""
    return tuple(
        OrderEntry(name=entry_name(name, design_hash), sequence=sequence)
        for name, design_hash, sequence in designs
    )
