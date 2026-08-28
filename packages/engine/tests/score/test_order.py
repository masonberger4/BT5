"""The vendor order file, checked against a real workbook.

Every plate assertion here reads the file back with openpyxl rather than
inspecting the writer's own idea of what it wrote. The failure this guards is a
workbook that opens cleanly and is rejected by the vendor's uploader, which is
discovered at the point of ordering and nowhere earlier.
"""

from __future__ import annotations

import io
from pathlib import Path

import pytest
from bt5.score.order import (
    DEFAULT_PLATE_NAME,
    IDT_HEADERS,
    PLATE_SIZE,
    OrderEntry,
    OrderError,
    entry_name,
    order_entries,
    plates,
    read_plate,
    sheet_name_of,
    wells,
    write_csv,
    write_idt_plate,
)

SEQ = "ATGAAACCCGGGTTTTAA"


def entries(n: int = 3) -> tuple[OrderEntry, ...]:
    return tuple(OrderEntry(f"design_{i}_abc{i:03d}", SEQ) for i in range(n))


class TestWells:
    def test_row_major_order_matching_the_template(self) -> None:
        w = wells()
        assert len(w) == PLATE_SIZE == 96
        assert w[0] == "A1"
        assert w[11] == "A12"
        assert w[12] == "B1", "the template fills across a row before moving down"
        assert w[-1] == "H12"

    def test_wells_are_unique(self) -> None:
        assert len(set(wells())) == PLATE_SIZE

    def test_asking_for_more_than_a_plate_is_refused(self) -> None:
        with pytest.raises(OrderError, match="a plate holds"):
            wells(97)


class TestOrderEntry:
    def test_rejects_anything_but_bare_uppercase_acgt(self) -> None:
        """A lowercase base or an IUPAC code reaches the synthesiser as an order
        rather than as a question."""
        for bad in ("atgaaa", "ATGNNN", "ATG AAA", "ATGU"):
            with pytest.raises(OrderError, match="bare uppercase ACGT"):
                OrderEntry("x_abc", bad)

    def test_rejects_an_empty_name_or_sequence(self) -> None:
        with pytest.raises(OrderError, match="needs a name"):
            OrderEntry("   ", SEQ)
        with pytest.raises(OrderError, match="empty sequence"):
            OrderEntry("x_abc", "")

    def test_the_tube_label_carries_the_design_hash(self) -> None:
        assert entry_name("cd19_car", "a1b2c3d4e5f6") == "cd19_car_a1b2c3d4e5f6"

    def test_a_label_without_a_hash_is_refused(self) -> None:
        """Two runs producing two sequences under one name is how a lab ends up
        with two tubes and an irreproducible result."""
        with pytest.raises(OrderError, match="not traceable"):
            entry_name("cd19_car", "")

    def test_order_entries_builds_labels_from_triples(self) -> None:
        built = order_entries([("cd19_car", "a1b2c3", SEQ)])
        assert built[0].name == "cd19_car_a1b2c3"
        assert built[0].length_bp == len(SEQ)


class TestPlateWorkbook:
    def test_the_sheet_name_is_the_plate_name(self) -> None:
        """In this template plate identity travels as sheet metadata, not as a
        column, so writing it anywhere else loses it."""
        data = write_idt_plate(entries(), plate_name="BT5 run 12")
        assert sheet_name_of(io.BytesIO(data)) == "BT5 run 12"

    def test_the_default_sheet_name_matches_the_shipped_template(self) -> None:
        assert sheet_name_of(io.BytesIO(write_idt_plate(entries()))) == DEFAULT_PLATE_NAME

    def test_exactly_three_columns_with_the_template_headers(self) -> None:
        from openpyxl import load_workbook

        sheet = load_workbook(io.BytesIO(write_idt_plate(entries()))).active
        header = next(sheet.iter_rows(values_only=True))
        assert header == IDT_HEADERS
        assert sheet.max_column == 3

    def test_all_96_well_positions_are_present_even_when_unused(self) -> None:
        rows = read_plate(io.BytesIO(write_idt_plate(entries(3))))
        assert len(rows) == PLATE_SIZE
        assert [r[0] for r in rows] == list(wells())

    def test_designs_land_in_row_major_order_and_unused_wells_are_blank(self) -> None:
        rows = read_plate(io.BytesIO(write_idt_plate(entries(3))))
        assert rows[0][1:] == ("design_0_abc000", SEQ)
        assert rows[1][0] == "A2"
        assert rows[2][1] == "design_2_abc002"
        assert rows[3] == ("A4", "", ""), "unused wells keep their position and nothing else"

    def test_a_full_plate_fills_every_well(self) -> None:
        rows = read_plate(io.BytesIO(write_idt_plate(entries(PLATE_SIZE))))
        assert all(name for _, name, _ in rows)
        assert rows[-1][0] == "H12"

    def test_more_than_one_plate_of_designs_is_refused_not_truncated(self) -> None:
        with pytest.raises(OrderError, match="do not fit one"):
            write_idt_plate(entries(PLATE_SIZE + 1))

    def test_plates_splits_rather_than_dropping_designs(self) -> None:
        chunks = plates(entries(PLATE_SIZE + 5))
        assert [len(c) for c in chunks] == [PLATE_SIZE, 5]
        assert chunks[1][0].name == entries(PLATE_SIZE + 5)[PLATE_SIZE].name

    def test_duplicate_names_are_refused(self) -> None:
        dupes = (OrderEntry("same_abc", SEQ), OrderEntry("same_abc", "ATGCCCTAA"))
        with pytest.raises(OrderError, match="duplicate order names"):
            write_idt_plate(dupes)

    def test_an_empty_order_is_refused(self) -> None:
        with pytest.raises(OrderError, match="nothing to order"):
            write_idt_plate(())

    @pytest.mark.parametrize("bad", ["run/12", "run:12", "run[1]", "run?", "a" * 32])
    def test_a_plate_name_excel_would_mangle_is_refused(self, bad: str) -> None:
        """Silently rewriting it would send the vendor a different plate name
        than the one on the bench."""
        with pytest.raises(OrderError):
            write_idt_plate(entries(), plate_name=bad)

    def test_writes_to_a_path(self, tmp_path: Path) -> None:
        target = tmp_path / "plate.xlsx"
        write_idt_plate(entries(), target, plate_name="run")
        assert target.exists()
        assert sheet_name_of(target) == "run"
        assert len(read_plate(target)) == PLATE_SIZE

    def test_the_sequence_survives_the_round_trip_exactly(self) -> None:
        """openpyxl will happily coerce a cell; a mangled base is a wrong order."""
        rows = read_plate(io.BytesIO(write_idt_plate(entries(1))))
        assert rows[0][2] == SEQ


class TestCsv:
    def test_two_columns_and_a_header(self) -> None:
        text = write_csv(entries(2))
        lines = text.strip().splitlines()
        assert lines[0] == "Name,Sequence"
        assert lines[1] == f"design_0_abc000,{SEQ}"
        assert len(lines) == 3

    def test_duplicate_names_are_refused_here_too(self) -> None:
        with pytest.raises(OrderError, match="duplicate order names"):
            write_csv((OrderEntry("same_abc", SEQ), OrderEntry("same_abc", SEQ)))

    def test_writes_to_a_path_without_doubling_line_endings(self, tmp_path: Path) -> None:
        target = tmp_path / "order.csv"
        write_csv(entries(2), target)
        raw = target.read_bytes()
        assert b"\r\r\n" not in raw, "newline='' keeps Windows endings from doubling"
        assert raw.decode().startswith("Name,Sequence\r\n")
